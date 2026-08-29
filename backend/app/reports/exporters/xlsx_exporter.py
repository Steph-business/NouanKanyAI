"""
app/reports/exporters/xlsx_exporter.py — Exportateur de rapports énergétiques au format XLSX (Excel).
"""

import io
import logging
from typing import Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.reports.exporters.base import BaseReportExporter
from app.reports.models import EnergyReportData

logger = logging.getLogger("nouankany.reports")

HEADER_FILL = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
ZEBRA_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="0F172A")
BOLD_FONT = Font(name="Arial", size=10, bold=True)
REGULAR_FONT = Font(name="Arial", size=10)

THIN_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)


class XLSXReportExporter(BaseReportExporter):
    """
    Générateur de classeurs Excel structurés et multi-onglets via openpyxl.
    """

    def export(
        self,
        report_data: EnergyReportData,
        chart_images: Optional[Dict[str, bytes]] = None,
    ) -> bytes:
        """
        Génère le classeur Excel (.xlsx).
        """
        wb = Workbook()

        # =============================================================
        # Feuille 1 : Synthèse & KPIs
        # =============================================================
        ws_kpi = wb.active
        ws_kpi.title = "Synthèse & KPIs"
        ws_kpi.views.sheetView[0].showGridLines = True

        ws_kpi["A1"] = f"NOUANKANY.AI — {report_data.title.upper()}"
        ws_kpi["A1"].font = TITLE_FONT
        ws_kpi["A2"] = f"Site : {report_data.site_name} | Période : {report_data.period_start} au {report_data.period_end} | Émis : {report_data.generated_at}"
        ws_kpi["A2"].font = Font(name="Arial", size=9, italic=True, color="64748B")

        # Résumé exécutif
        ws_kpi["A4"] = "Résumé Exécutif :"
        ws_kpi["A4"].font = BOLD_FONT
        ws_kpi["A5"] = report_data.executive_summary
        ws_kpi["A5"].font = REGULAR_FONT

        # Tableau des KPIs
        ws_kpi["A7"] = "INDICATEUR DE PERFORMANCE (KPI)"
        ws_kpi["B7"] = "VALEUR"
        ws_kpi["C7"] = "UNITÉ"
        for col in ["A", "B", "C"]:
            cell = ws_kpi[f"{col}7"]
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center")

        k = report_data.kpis
        kpi_rows = [
            ("Consommation Totale", k.total_energy_kwh, "kWh"),
            ("Facture Électrique Globale", k.total_cost_fcfa, "FCFA"),
            ("Consommation Heures de Pointe (19h-23h)", k.peak_hours_energy_kwh, "kWh"),
            ("Coût Heures de Pointe", k.peak_hours_cost_fcfa, "FCFA"),
            ("Économies Réalisées par Effacement", k.peak_shaving_savings_fcfa, "FCFA"),
            ("Facteur de Puissance Moyen (cos φ)", k.average_power_factor_cos_phi, "sans dim."),
            ("Puissance Maximale Atteinte", k.max_peak_power_kw, "kW"),
            ("Puissance Souscrite CIE", k.subscribed_power_limit_kw, "kW"),
            ("Émissions de CO2 Estimées", k.carbon_emissions_kg_co2, "kg CO2"),
            ("Nombre d'Anomalies Détectées", k.anomaly_incidents_count, "incidents"),
        ]

        for idx, (label, val, unit) in enumerate(kpi_rows, start=8):
            ws_kpi[f"A{idx}"] = label
            ws_kpi[f"B{idx}"] = val
            ws_kpi[f"C{idx}"] = unit

            ws_kpi[f"A{idx}"].font = BOLD_FONT if "Globale" in label or "Totale" in label else REGULAR_FONT
            ws_kpi[f"B{idx}"].font = BOLD_FONT
            ws_kpi[f"C{idx}"].font = REGULAR_FONT

            for c in ["A", "B", "C"]:
                ws_kpi[f"{c}{idx}"].border = THIN_BORDER
                if idx % 2 == 1:
                    ws_kpi[f"{c}{idx}"].fill = ZEBRA_FILL

            # Format numérique
            if unit == "FCFA":
                ws_kpi[f"B{idx}"].number_format = "#,##0"
            elif unit == "kWh" or unit == "kW":
                ws_kpi[f"B{idx}"].number_format = "#,##0.0"

        # =============================================================
        # Feuille 2 : Détail des Équipements
        # =============================================================
        ws_mach = wb.create_sheet(title="Consommation Équipements")
        ws_mach.views.sheetView[0].showGridLines = True

        mach_headers = ["Nom de la Machine", "Catégorie", "Énergie (kWh)", "Coût (FCFA)", "Heures de Marche (h)", "Statut", "Note Efficacité"]
        for c_idx, h in enumerate(mach_headers, start=1):
            col_letter = get_column_letter(c_idx)
            cell = ws_mach[f"{col_letter}1"]
            cell.value = h
            cell.fill = SUBHEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for r_idx, m in enumerate(report_data.machines, start=2):
            ws_mach[f"A{r_idx}"] = m.machine_name
            ws_mach[f"B{r_idx}"] = m.category
            ws_mach[f"C{r_idx}"] = m.energy_kwh
            ws_mach[f"D{r_idx}"] = m.cost_fcfa
            ws_mach[f"E{r_idx}"] = m.running_hours
            ws_mach[f"F{r_idx}"] = m.status
            ws_mach[f"G{r_idx}"] = m.efficiency_grade

            ws_mach[f"C{r_idx}"].number_format = "#,##0.0"
            ws_mach[f"D{r_idx}"].number_format = "#,##0"

            for c_idx in range(1, 8):
                cell = ws_mach[f"{get_column_letter(c_idx)}{r_idx}"]
                cell.border = THIN_BORDER
                cell.font = REGULAR_FONT
                if r_idx % 2 == 1:
                    cell.fill = ZEBRA_FILL

        # =============================================================
        # Feuille 3 : Registre des Anomalies
        # =============================================================
        ws_anom = wb.create_sheet(title="Anomalies & Dérives")
        ws_anom.views.sheetView[0].showGridLines = True

        anom_headers = ["Identifiant Incident", "Horodatage", "Équipement", "Sévérité", "Score Anomalie", "Description du Dysfonctionnement", "Action Corrective"]
        for c_idx, h in enumerate(anom_headers, start=1):
            col_letter = get_column_letter(c_idx)
            cell = ws_anom[f"{col_letter}1"]
            cell.value = h
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

        for r_idx, a in enumerate(report_data.anomalies, start=2):
            ws_anom[f"A{r_idx}"] = a.incident_id
            ws_anom[f"B{r_idx}"] = a.timestamp
            ws_anom[f"C{r_idx}"] = a.equipment
            ws_anom[f"D{r_idx}"] = a.severity.upper()
            ws_anom[f"E{r_idx}"] = a.anomaly_score
            ws_anom[f"F{r_idx}"] = a.description
            ws_anom[f"G{r_idx}"] = a.action_taken

            for c_idx in range(1, 8):
                cell = ws_anom[f"{get_column_letter(c_idx)}{r_idx}"]
                cell.border = THIN_BORDER
                cell.font = REGULAR_FONT

        # =============================================================
        # Feuille 4 : Recommandations IA & ROI
        # =============================================================
        ws_rec = wb.create_sheet(title="Recommandations IA")
        ws_rec.views.sheetView[0].showGridLines = True

        rec_headers = ["Priorité", "Titre de l'Action", "Cible", "Description Détaillée", "Gain Estimé (FCFA)", "Gain Estimé (kWh)"]
        for c_idx, h in enumerate(rec_headers, start=1):
            col_letter = get_column_letter(c_idx)
            cell = ws_rec[f"{col_letter}1"]
            cell.value = h
            cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
            cell.font = HEADER_FONT

        for r_idx, rec in enumerate(report_data.recommendations, start=2):
            ws_rec[f"A{r_idx}"] = f"P{rec.priority}"
            ws_rec[f"B{r_idx}"] = rec.title
            ws_rec[f"C{r_idx}"] = rec.target_equipment
            ws_rec[f"D{r_idx}"] = rec.description
            ws_rec[f"E{r_idx}"] = rec.estimated_savings_fcfa
            ws_rec[f"F{r_idx}"] = rec.estimated_savings_kwh

            ws_rec[f"E{r_idx}"].number_format = "#,##0"
            ws_rec[f"F{r_idx}"].number_format = "#,##0"

            for c_idx in range(1, 7):
                cell = ws_rec[f"{get_column_letter(c_idx)}{r_idx}"]
                cell.border = THIN_BORDER
                cell.font = REGULAR_FONT

        # Ajustement automatique de la largeur des colonnes sur toutes les feuilles
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        lines = str(cell.value).split("\n")
                        max_len = max(max_len, max(len(l) for l in lines))
                sheet.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()
